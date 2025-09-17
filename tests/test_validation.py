from io import BytesIO
from importlib import util
from pathlib import Path
import asyncio
from openpyxl import Workbook
from starlette.datastructures import UploadFile

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SERVER_PATH = PROJECT_ROOT / "server.py"
SPEC = util.spec_from_file_location("server", SERVER_PATH)
assert SPEC and SPEC.loader
server = util.module_from_spec(SPEC)
SPEC.loader.exec_module(server)  # type: ignore[arg-type]


def _build_workbook_missing_sheet1_column() -> bytes:
    workbook = Workbook()

    sheet1 = workbook.active
    sheet1.title = "RawData1"
    sheet1["A1"] = "header"
    # Ensure sheet 1 has columns up to AL but not BZ
    sheet1.cell(row=1, column=38, value="AL header")
    sheet1["A2"] = "data"

    sheet2 = workbook.create_sheet("RawData2")
    sheet2["A1"] = "header"
    sheet2.cell(row=1, column=48, value="AV header")
    sheet2["A2"] = "data"

    sheet3 = workbook.create_sheet("RawData3")
    sheet3["A1"] = "header"
    sheet3.cell(row=1, column=92, value="CN header")
    sheet3["A2"] = "data"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def test_missing_required_column_returns_validation_error():
    file_bytes = _build_workbook_missing_sheet1_column()
    upload = UploadFile(filename="test.xlsx", file=BytesIO(file_bytes))

    response = asyncio.run(
        server.process_excel(
            file=upload,
            existing_file=None,
            output_sheet_name="Q1-Q2-Q3-Q4-2024",
            raw_sheet1_name="",
            raw_sheet2_name="",
            raw_sheet3_name="",
            deal_column_name="N",
        )
    )

    assert isinstance(response, dict)
    assert "error" in response
    assert "missing required column BZ" in response["error"]
