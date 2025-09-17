from fastapi import FastAPI, UploadFile, Form, File
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from io import BytesIO
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Set, Optional

app = FastAPI(title="VARO REBILLING Excel Processor", version="1.0.0")

# Add CORS middleware for local development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files (for production deployment)
import os
static_dir = os.path.dirname(os.path.abspath(__file__))
app.mount("/static", StaticFiles(directory=static_dir), name="static")

# Serve frontend assets at root paths for production deployment
@app.get("/styles.css")
async def styles_css():
    return FileResponse(os.path.join(static_dir, 'styles.css'))

@app.get("/app.js")
async def app_js():
    return FileResponse(os.path.join(static_dir, 'app.js'))

class ExcelProcessor:
    def __init__(self):
        self.workbook = None
        self.lookup_tables = {}

    def process_excel_file(
        self,
        file_data: bytes,
        settings: Dict[str, Any],
        existing_file_data: Optional[bytes] = None
    ) -> BytesIO:
        """Main processing function that handles both steps and formatting"""

        # Load the input workbook
        self.workbook = load_workbook(BytesIO(file_data), data_only=True)

        # Validate the file structure
        self.validate_file(settings)

        # Build lookup tables for Step 2
        self.build_lookup_tables(settings)

        # Step 1: Build formatted report with column mapping
        new_workbook = self.build_step1_report(settings)

        # Step 2: Enrich data using business rules
        self.enrich_step2_data(new_workbook, settings)

        # Apply comprehensive Excel formatting
        self.apply_comprehensive_formatting(new_workbook, settings)

        # Highlight differences against an existing formatted workbook if provided
        if existing_file_data:
            self.highlight_differences(new_workbook, existing_file_data, settings)

        # Save to BytesIO
        output = BytesIO()
        new_workbook.save(output)
        output.seek(0)

        return output

    def validate_file(self, settings: Dict[str, Any]) -> None:
        """Validate that required sheets and columns exist"""
        sheet_names = self.workbook.sheetnames

        raw_sheet1 = settings.get('raw_sheet1_name') or sheet_names[0]
        raw_sheet2 = settings.get('raw_sheet2_name') or sheet_names[1]
        raw_sheet3 = settings.get('raw_sheet3_name') or sheet_names[2]

        if raw_sheet1 not in sheet_names:
            raise ValueError(f"Raw Sheet 1 '{raw_sheet1}' not found")
        if raw_sheet2 not in sheet_names:
            raise ValueError(f"Raw Sheet 2 '{raw_sheet2}' not found")
        if raw_sheet3 not in sheet_names:
            raise ValueError(f"Raw Sheet 3 '{raw_sheet3}' not found")

        # Verify essential data exists
        sheet1 = self.workbook[raw_sheet1]
        if sheet1.max_row < 2:
            raise ValueError("Raw Sheet 1 must have at least one data row")

    def build_step1_report(self, settings: Dict[str, Any]) -> Workbook:
        """Step 1: Create formatted report with proper column mapping"""

        new_workbook = Workbook()
        ws = new_workbook.active
        ws.title = settings.get('output_sheet_name', 'Q1-Q2-Q3-Q4-2024')

        # Headers A-V
        headers = [
            'Varo deal', 'VSA deal', 'VESSEL', 'VMAG %', 'L/C costs',
            'Load insp', 'Discharge inspection', 'Superintendent', 'CIN insurance',
            'CLI insurance', 'Provisional charge', 'TOTAL USD', 'VARO comments',
            'Product', 'Hedge', 'Qty BBL', 'Inco', 'Contractual Location',
            'Risk', 'Date', 'VSA comments', 'Additional information'
        ]

        # Add headers to row 1
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Add 1 blank row after headers (row 2)

        # Get raw data from Sheet 1
        raw_sheet1_name = settings.get('raw_sheet1_name') or self.workbook.sheetnames[0]
        raw_sheet1 = self.workbook[raw_sheet1_name]

        # Extract data rows (skip header row)
        raw_data = []
        for row_idx, row in enumerate(raw_sheet1.iter_rows(min_row=2, values_only=True), 2):
            if row and any(cell for cell in row):  # Skip completely empty rows
                raw_data.append(row)

        # Map raw data to new format
        mapped_data = []
        for row in raw_data:
            # Ensure row has enough columns
            padded_row = list(row) + [None] * (100 - len(row))

            new_row = [None] * 22  # A-V = 22 columns

            # Column mapping: B←F, C←AA, N←M, O←L, P←Q, Q←AB, R←AD, S←AL, T←AM
            new_row[1] = padded_row[5]    # B ← F (column F = index 5)
            new_row[2] = padded_row[26]   # C ← AA (column AA = index 26)
            new_row[13] = padded_row[12]  # N ← M (column M = index 12)
            new_row[14] = padded_row[11]  # O ← L (column L = index 11)
            new_row[15] = padded_row[16]  # P ← Q (column Q = index 16)
            new_row[16] = padded_row[27]  # Q ← AB (column AB = index 27)
            new_row[17] = padded_row[29]  # R ← AD (column AD = index 29)
            new_row[18] = padded_row[37]  # S ← AL (column AL = index 37)
            new_row[19] = self.parse_date(padded_row[38])  # T ← AM (column AM = index 38)

            mapped_data.append(new_row)

        # Sort by date (column T, index 19)
        def date_sort_key(row):
            date_val = self.parse_date(row[19])
            return date_val if date_val else datetime.max

        mapped_data.sort(key=date_sort_key)

        # Group by month and add to worksheet
        current_row = 3  # Start at row 3 (after header and blank row)
        current_month = None
        current_year = None

        for row_data in mapped_data:
            date_val = self.parse_date(row_data[19])

            if date_val:
                month = date_val.month
                year = date_val.year

                if current_month != month or current_year != year:
                    # Add spacing before new month (except first)
                    if current_month is not None:
                        current_row += 3  # 3 blank rows between months

                    # Add month header (e.g., JAN-24)
                    month_name = date_val.strftime('%b').upper()
                    year_short = str(year)[-2:]
                    month_header = f"{month_name}-{year_short}"

                    ws.cell(row=current_row, column=1, value=month_header)
                    current_row += 1

                    current_month = month
                    current_year = year

            # Format date for display as DD/MM/YYYY
            if row_data[19] and isinstance(row_data[19], datetime):
                row_data[19] = row_data[19].strftime('%d/%m/%Y')

            # Add data row
            for col, value in enumerate(row_data, 1):
                if value is not None:
                    ws.cell(row=current_row, column=col, value=value)

            current_row += 1

        return new_workbook

    def enrich_step2_data(self, workbook: Workbook, settings: Dict[str, Any]) -> None:
        """Step 2: Apply all 8 business rules"""

        ws = workbook.active
        locks = set()  # Track locked cells

        # Process each data row
        for row_idx in range(3, ws.max_row + 1):  # Start from row 3 (after header + blank)

            # Get row values
            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 23)]

            # Skip empty rows and month header rows
            if not row_values[1] and not row_values[13] and not row_values[14]:  # No VSA deal, product, or hedge
                continue

            deal = self.normalize(row_values[1])     # VSA deal (B)
            product = self.normalize(row_values[13]) # Product (N)
            hedge = self.normalize(row_values[14])   # Hedge (O)

            # Rule 1: MIDLANDS product - set columns E-K to 0 and lock
            if product == 'MIDLANDS':
                for col in range(5, 12):  # E-K (columns 5-11)
                    ws.cell(row=row_idx, column=col, value=0)
                    locks.add(f"{row_idx},{col}")

            # Rule 2: WHB+CIF deals - set insurance columns I,J to 0 and lock
            if deal and deal in self.lookup_tables.get('whb_cif_deals', set()):
                if f"{row_idx},9" not in locks:  # Column I (index 9)
                    ws.cell(row=row_idx, column=9, value=0)
                    locks.add(f"{row_idx},9")
                if f"{row_idx},10" not in locks:  # Column J (index 10)
                    ws.cell(row=row_idx, column=10, value=0)
                    locks.add(f"{row_idx},10")

            # Rule 3: LC Costs (E) - BOT + BLC totals
            if f"{row_idx},5" not in locks and ws.cell(row=row_idx, column=5).value != 0:
                bot_cost = self.lookup_tables.get('costs_map', {}).get(f"{deal},BOT", 0)
                blc_cost = self.lookup_tables.get('costs_map', {}).get(f"{deal},BLC", 0)
                total_cost = bot_cost + blc_cost
                if total_cost != 0:
                    ws.cell(row=row_idx, column=5, value=total_cost)

            # Rule 4: CIN insurance (I) - CIN costs
            if f"{row_idx},9" not in locks and ws.cell(row=row_idx, column=9).value != 0:
                cin_cost = self.lookup_tables.get('costs_map', {}).get(f"{deal},CIN", 0)
                if cin_cost != 0:
                    ws.cell(row=row_idx, column=9, value=cin_cost)

            # Rule 5: CLI insurance (J) - CLI costs
            if f"{row_idx},10" not in locks and ws.cell(row=row_idx, column=10).value != 0:
                cli_cost = self.lookup_tables.get('costs_map', {}).get(f"{deal},CLI", 0)
                if cli_cost != 0:
                    ws.cell(row=row_idx, column=10, value=cli_cost)

            # Rule 9: INS/INQ/INA insurance (F) - Load inspection costs
            if f"{row_idx},6" not in locks and ws.cell(row=row_idx, column=6).value != 0:
                ins_cost = self.lookup_tables.get('costs_map', {}).get(f"{deal},INS", 0)
                inq_cost = self.lookup_tables.get('costs_map', {}).get(f"{deal},INQ", 0)
                ina_cost = self.lookup_tables.get('costs_map', {}).get(f"{deal},INA", 0)
                total_inspection_cost = ins_cost + inq_cost + ina_cost
                if total_inspection_cost != 0:
                    ws.cell(row=row_idx, column=6, value=total_inspection_cost)

            # Rule 6: TOTAL calculation (L) - SUM(E:K)
            total = 0
            for col in range(5, 12):  # E-K (columns 5-11)
                val = ws.cell(row=row_idx, column=col).value
                if val and isinstance(val, (int, float)):
                    total += val
            ws.cell(row=row_idx, column=12, value=total)  # Column L

            # Rule 7: VSA comments (U) - from hedge lookup
            if hedge and hedge in self.lookup_tables.get('hedge_to_br', {}):
                br_value = self.lookup_tables['hedge_to_br'][hedge]
                ws.cell(row=row_idx, column=21, value=br_value)  # Column U

            # Rule 8: Additional information (V) - from hedge lookup
            if hedge and hedge in self.lookup_tables.get('hedge_to_cn', {}):
                cn_value = self.lookup_tables['hedge_to_cn'][hedge]
                ws.cell(row=row_idx, column=22, value=cn_value)  # Column V

    def build_lookup_tables(self, settings: Dict[str, Any]) -> None:
        """Build lookup tables for Step 2 business rules"""

        raw_sheet1_name = settings.get('raw_sheet1_name') or self.workbook.sheetnames[0]
        raw_sheet2_name = settings.get('raw_sheet2_name') or self.workbook.sheetnames[1]
        raw_sheet3_name = settings.get('raw_sheet3_name') or self.workbook.sheetnames[2]

        # WHB+CIF deals from Sheet 1
        whb_cif_deals = set()
        sheet1 = self.workbook[raw_sheet1_name]
        for row in sheet1.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 77:
                # Check if AB=CIF (index 27) and BZ=WHB (index 77)
                if len(row) > 27 and len(row) > 77 and row[27] == 'CIF' and row[77] == 'WHB':
                    deal = self.normalize(row[5])  # Column F (index 5)
                    if deal:
                        whb_cif_deals.add(deal)

        # Costs map from Sheet 2
        costs_map = {}
        sheet2 = self.workbook[raw_sheet2_name]
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 47:
                deal = self.normalize(row[13])  # Column N (index 13)
                cost_type = self.normalize(row[42])  # Column AQ (index 42)
                amount = self.safe_float(row[47])  # Column AV (index 47)

                if deal and cost_type:
                    key = f"{deal},{cost_type}"
                    costs_map[key] = costs_map.get(key, 0) + amount

        # Hedge maps from Sheet 3
        hedge_to_br = {}
        hedge_to_cn = {}
        sheet3 = self.workbook[raw_sheet3_name]
        for row in sheet3.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 91:
                hedge = self.normalize(row[12])  # Column M (index 12)
                br_value = row[69] if len(row) > 69 else None  # Column BR (index 69)
                cn_value = row[91] if len(row) > 91 else None  # Column CN (index 91)

                if hedge:
                    if br_value and hedge not in hedge_to_br:
                        hedge_to_br[hedge] = br_value
                    if cn_value and hedge not in hedge_to_cn:
                        hedge_to_cn[hedge] = cn_value

        self.lookup_tables = {
            'whb_cif_deals': whb_cif_deals,
            'costs_map': costs_map,
            'hedge_to_br': hedge_to_br,
            'hedge_to_cn': hedge_to_cn
        }

    def apply_comprehensive_formatting(self, workbook: Workbook, settings: Dict[str, Any]) -> None:
        """Apply all Excel formatting requirements"""

        ws = workbook.active

        # Define styles
        header_font = Font(bold=True, size=14, name='Arial')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        month_font = Font(bold=True, size=12, name='Arial')
        month_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        month_alignment = Alignment(horizontal='center', vertical='center')

        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Column L special borders (TOTAL USD) - only vertical lines
        column_l_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000')
        )
        # Removed column_l_fill - no background for column L

        # Row heights
        ws.row_dimensions[1].height = 85   # Header row - 85pt

        # Set all other rows to height 15
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 15

        # No freeze panes

        # Column widths (auto-fit based on content)
        column_widths = [18, 20, 22, 12, 14, 16, 22, 18, 16, 16, 20, 16, 30, 18, 14, 14, 14, 24, 12, 14, 30, 30]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Apply formatting to all cells
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, 23):  # A-V
                cell = ws.cell(row=row_idx, column=col_idx)

                # Column L (TOTAL USD) - vertical borders for ALL rows including header, no background
                if col_idx == 12:
                    cell.border = column_l_border
                    cell.alignment = center_alignment if row_idx > 1 else header_alignment
                    # Special formatting for column L header
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=14, name='Arial')

                # Header row formatting (row 1) - no gray background
                elif row_idx == 1:
                    cell.font = header_font
                    cell.alignment = header_alignment

                # Month cells formatting (column A, rows 3+)
                elif col_idx == 1 and row_idx >= 3 and cell.value:
                    cell_text = str(cell.value).strip().upper()
                    # Check for month pattern (JAN-24, FEB-24, etc.)
                    if any(month in cell_text for month in ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                                                           'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']):
                        cell.font = month_font
                        cell.fill = month_fill
                        cell.alignment = month_alignment

                # Data cells formatting (rows 3+)
                elif row_idx >= 3 and cell.value is not None:
                    # Columns U and V - left align, no borders
                    if col_idx in [21, 22]:  # U, V
                        cell.alignment = left_alignment
                    # All other data cells - center align, no borders
                    else:
                        cell.alignment = center_alignment

    def highlight_differences(
        self,
        workbook: Workbook,
        existing_file_data: bytes,
        settings: Dict[str, Any]
    ) -> None:
        """Compare generated workbook with an existing formatted workbook and highlight deltas."""

        try:
            existing_wb = load_workbook(BytesIO(existing_file_data), data_only=True)
        except Exception:
            # If the formatted workbook cannot be read, skip highlighting
            return

        output_sheet_name = settings.get('output_sheet_name') or workbook.active.title

        if output_sheet_name in existing_wb.sheetnames:
            existing_ws = existing_wb[output_sheet_name]
        else:
            existing_ws = existing_wb.active

        existing_rows = self.extract_existing_rows(existing_ws)
        ws = workbook.active

        seen_deals: Set[str] = set()

        for row_idx in range(1, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 23)]
            deal_key = self.normalize(row_values[1])

            if not deal_key or deal_key == 'VSA DEAL':
                continue

            if deal_key in existing_rows:
                existing_values = existing_rows[deal_key]
                seen_deals.add(deal_key)

                for col_idx in range(1, 23):
                    new_value = row_values[col_idx - 1]
                    old_value = existing_values[col_idx - 1]

                    if self.values_differ(new_value, old_value):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        self.mark_cell_red(cell)
            else:
                # Entire deal is new – mark populated cells
                for col_idx in range(1, 23):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if not self.is_blank(cell.value):
                        self.mark_cell_red(cell)

        missing_deals = sorted(set(existing_rows.keys()) - seen_deals)

        if missing_deals:
            discrepancy_ws = workbook.create_sheet(title='Missing from Raw')
            headers = ['VSA deal', 'Product', 'Qty BBL', 'Notes']
            discrepancy_ws.append(headers)

            header_font = Font(bold=True, size=12, name='Arial')
            for col_idx in range(1, len(headers) + 1):
                discrepancy_ws.cell(row=1, column=col_idx).font = header_font

            for row_offset, deal in enumerate(missing_deals, start=2):
                existing_values = existing_rows[deal]
                product = existing_values[13] if len(existing_values) > 13 else None
                qty = existing_values[15] if len(existing_values) > 15 else None

                discrepancy_ws.cell(row=row_offset, column=1, value=existing_values[1])
                discrepancy_ws.cell(row=row_offset, column=2, value=product)
                discrepancy_ws.cell(row=row_offset, column=3, value=qty)
                discrepancy_ws.cell(row=row_offset, column=4, value='Not present in latest raw data')

                for col_idx in range(1, 4):
                    cell = discrepancy_ws.cell(row=row_offset, column=col_idx)
                    if not self.is_blank(cell.value):
                        self.mark_cell_red(cell)

            # Set simple column widths for readability
            column_widths = [18, 18, 14, 40]
            for idx, width in enumerate(column_widths, 1):
                discrepancy_ws.column_dimensions[get_column_letter(idx)].width = width

    def extract_existing_rows(self, worksheet) -> Dict[str, List[Any]]:
        """Extract existing rows keyed by deal identifier from a formatted worksheet."""

        data: Dict[str, List[Any]] = {}

        for row_idx in range(2, worksheet.max_row + 1):
            deal_value = worksheet.cell(row=row_idx, column=2).value

            if self.is_blank(deal_value):
                continue

            deal_key = self.normalize(deal_value)

            if not deal_key or deal_key == 'VSA DEAL':
                continue

            row_values = [worksheet.cell(row=row_idx, column=col).value for col in range(1, 23)]
            data[deal_key] = row_values

        return data

    def is_blank(self, value) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and value.strip() == '':
            return True
        return False

    def values_differ(self, new_value, old_value) -> bool:
        if self.is_blank(new_value) and self.is_blank(old_value):
            return False

        # Normalize dates
        if isinstance(new_value, datetime):
            new_value = new_value.strftime('%d/%m/%Y')
        if isinstance(old_value, datetime):
            old_value = old_value.strftime('%d/%m/%Y')

        # Strip strings for comparison
        if isinstance(new_value, str):
            new_value = new_value.strip()
        if isinstance(old_value, str):
            old_value = old_value.strip()

        # Compare numerics with tolerance
        try:
            new_float = float(new_value)
            old_float = float(old_value)
            return abs(new_float - old_float) > 0.0001
        except (TypeError, ValueError):
            return new_value != old_value

    def mark_cell_red(self, cell) -> None:
        """Apply a red font color to highlight changes without altering other attributes."""

        font = cell.font or Font(name='Arial')
        cell.font = font.copy(color='00FF0000')

    def normalize(self, value) -> str:
        """Normalize values for consistent comparison"""
        return str(value).strip().upper() if value else ''

    def parse_date(self, value) -> Optional[datetime]:
        """Parse various date formats"""
        if not value:
            return None

        if isinstance(value, datetime):
            return value

        # Try parsing Excel date number
        if isinstance(value, (int, float)):
            try:
                # Excel date serial number (days since 1900-01-01)
                from datetime import date, timedelta
                base_date = date(1899, 12, 30)  # Excel epoch
                return datetime.combine(base_date + timedelta(days=value), datetime.min.time())
            except:
                pass

        # Try parsing string date
        if isinstance(value, str):
            try:
                return datetime.strptime(value, '%Y-%m-%d')
            except:
                try:
                    return datetime.strptime(value, '%d/%m/%Y')
                except:
                    pass

        return None

    def safe_float(self, value) -> float:
        """Safely convert value to float"""
        try:
            return float(value) if value else 0.0
        except (ValueError, TypeError):
            return 0.0

@app.post("/process")
async def process_excel(
    file: UploadFile,
    existing_file: Optional[UploadFile] = File(None),
    output_sheet_name: str = Form("Q1-Q2-Q3-Q4-2024"),
    raw_sheet1_name: str = Form(""),
    raw_sheet2_name: str = Form(""),
    raw_sheet3_name: str = Form(""),
    deal_column_name: str = Form("N")
):
    """Process Excel file with all business rules and formatting"""

    try:
        # Read uploaded file
        file_data = await file.read()

        # Prepare settings
        settings = {
            'output_sheet_name': output_sheet_name,
            'raw_sheet1_name': raw_sheet1_name if raw_sheet1_name else None,
            'raw_sheet2_name': raw_sheet2_name if raw_sheet2_name else None,
            'raw_sheet3_name': raw_sheet3_name if raw_sheet3_name else None,
            'deal_column_name': deal_column_name
        }

        existing_data = await existing_file.read() if existing_file else None

        # Process the Excel file
        processor = ExcelProcessor()
        output_buffer = processor.process_excel_file(file_data, settings, existing_data)

        # Return formatted Excel file
        headers = {
            "Content-Disposition": "attachment; filename=formatted_output.xlsx"
        }

        return StreamingResponse(
            BytesIO(output_buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except Exception as e:
        return {"error": str(e)}

@app.get("/health")
async def health_check():
    """Health check endpoint for Railway deployment"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

@app.get("/")
async def read_index():
    """Serve the main dashboard page"""
    return FileResponse('index.html')

if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
